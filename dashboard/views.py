from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import logout
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from rest_framework import status, generics, serializers, viewsets, mixins
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate, login
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
from .serializers import UserSerializer, LoginSerializer, RegisterSerializer, DepartmentSerializer, SubDepartmentSerializer, DivisionBranchSerializer, BranchDepartmentLinkSerializer, LogoSerializer, DataEntryRecordSerializer, BranchSerializer, ActivityLogSerializer
from .models import User, Department, SubDepartment, DivisionBranch, BranchDepartmentLink, Logo, UserSubDepartment, DataEntryRecord, DataEntryFile, ActivityLog
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from rest_framework.decorators import api_view, action
from django.core.exceptions import PermissionDenied
import json
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
from django.http import Http404
import base64
from django.urls import path
from django.db import transaction
from datetime import datetime
from django.utils import timezone
from .signals import log_activity
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO
import tempfile
import zipfile
import os
import mimetypes

# Create your views here.

def login_view(request):
    # Get active logo and organization name
    try:
        active_logo = Logo.objects.filter(is_active=True).first()
        context = {
            'logo_data': base64.b64encode(active_logo.logo_data).decode('utf-8') if active_logo else None,
            'organization_name': active_logo.organization_name if active_logo else 'Organization Name'
        }
    except Exception as e:
        print(f"Error fetching logo: {str(e)}")
        context = {
            'logo_data': None,
            'organization_name': 'Organization Name'
        }
    return render(request, 'login.html', context)

@login_required
def logout_view(request):
    # Log the logout activity before logging out
    log_activity(
        user=request.user,
        action='logout',
        page='Logout',
        model_name='User',
        object_id=request.user.id,
        details={'username': request.user.username},
        request=request
    )
    
    # Perform the logout
    logout(request)
    return redirect('login_view')

@login_required
def dashboard_home(request):
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Dashboard Home',
        model_name='Dashboard',
        request=request
    )
    
    # Get counts for dashboard with detailed logging
    print("\n=== Database Query Results ===")
    
    # Users count
    users = User.objects.all()
    total_users = users.count()
    
    # Records count
    records = DataEntryRecord.objects.all()
    total_records = records.count()
    
    # Departments count
    departments = Department.objects.all()
    total_departments = departments.count()
    
    # Sub-Departments count
    sub_departments = SubDepartment.objects.all()
    total_sub_departments = sub_departments.count()
    
    # Branches count
    branches = DivisionBranch.objects.all()
    total_branches = branches.count()
    

    # Add user data and counts to the context
    context = {
        'user': request.user,
        'user_permissions': request.user.get_all_permissions(),
        'total_users': total_users,
        'total_records': total_records,
        'total_departments': total_departments,
        'total_sub_departments': total_sub_departments,
        'total_branches': total_branches,
    }
    return render(request, 'dashboard/home.html', context)

@login_required
def dashboard_config(request):
    return render(request, 'dashboard/config.html')

@login_required
def create_user(request):
    if not request.user.has_perm('dashboard.can_view_create_user'):
        return render(request, 'dashboard/403.html')
    return render(request, 'dashboard/create_user.html')

@login_required
def register(request):
    if not request.user.can_view_register:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Register',
        model_name='Register',
        request=request
    )
    
    # Get record counts for each sub-department with IDs
    record_counts = DataEntryRecord.objects.values(
        'branch__id',
        'branch__name',
        'department__id',
        'department__name',
        'sub_department__id',
        'sub_department__name'
    ).annotate(
        count=Count('id')
    ).order_by('branch__name', 'department__name', 'sub_department__name')
    
    # Organize the data by branch and department
    organized_data = {}
    for record in record_counts:
        branch_id = record['branch__id']
        branch_name = record['branch__name']
        dept_id = record['department__id']
        dept_name = record['department__name']
        sub_dept_id = record['sub_department__id']
        sub_dept_name = record['sub_department__name']
        count = record['count']
        
        if branch_id not in organized_data:
            organized_data[branch_id] = {
                'name': branch_name,
                'departments': {}
            }
        if dept_id not in organized_data[branch_id]['departments']:
            organized_data[branch_id]['departments'][dept_id] = {
                'name': dept_name,
                'sub_departments': {}
            }
        
        organized_data[branch_id]['departments'][dept_id]['sub_departments'][sub_dept_id] = {
            'name': sub_dept_name,
            'count': count
        }
    
    # Handle Excel download
    if request.GET.get('download') == 'excel':
        # Log the download action
        log_activity(
            user=request.user,
            action='download',
            page='Register',
            model_name='Register',
            details={'file_type': 'Excel', 'file_name': 'record_counts.xlsx'},
            request=request
        )
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Record Counts"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Define headers
        headers = [
            'Branch/Division',
            'Department',
            'Sub-Department',
            'Number of Records'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row = 2
        for branch_id, branch_data in organized_data.items():
            for dept_id, dept_data in branch_data['departments'].items():
                for sub_dept_id, sub_dept_data in dept_data['sub_departments'].items():
                    ws.cell(row=row, column=1, value=f"{branch_id} - {branch_data['name']}")
                    ws.cell(row=row, column=2, value=f"{dept_id} - {dept_data['name']}")
                    ws.cell(row=row, column=3, value=f"{sub_dept_id} - {sub_dept_data['name']}")
                    ws.cell(row=row, column=4, value=sub_dept_data['count'])
                    row += 1
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
        
        # Create the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=record_counts.xlsx'
        
        # Save the workbook to the response
        wb.save(response)
        return response
    
    context = {
        'organized_data': organized_data,
    }
    return render(request, 'dashboard/report/register.html', context)

# API Views
@method_decorator(csrf_exempt, name='dispatch')
class LoginView(generics.CreateAPIView):
    permission_classes = (AllowAny,)  # Allow any user to access login
    serializer_class = LoginSerializer

    def post(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            login_field = serializer.validated_data['username']  # This could be username or email
            password = serializer.validated_data['password']
            
            # Try to find user by email first
            try:
                user = User.objects.get(email=login_field)
            except User.DoesNotExist:
                # If not found by email, try username
                try:
                    user = User.objects.get(username=login_field)
                except User.DoesNotExist:
                    return Response(
                        {'error': 'Username or email is incorrect'},
                        status=status.HTTP_401_UNAUTHORIZED
                    )
            
            # Try to authenticate with email
            authenticated_user = authenticate(email=user.email, password=password)
            if not authenticated_user:
                # If email auth fails, try username
                authenticated_user = authenticate(username=user.username, password=password)
                if not authenticated_user:
                    return Response(
                        {'error': 'Password is incorrect'},
                        status=status.HTTP_401_UNAUTHORIZED
                    )
            
            # Check if user is active
            if authenticated_user.status != 'active':
                return Response(
                    {'error': 'Your account is inactive. Please contact administrator.'},
                    status=status.HTTP_401_UNAUTHORIZED
                )
            
            # If all checks pass, generate tokens
            refresh = RefreshToken.for_user(authenticated_user)
            
            # Set session authentication
            login(request, authenticated_user)
            
            # Return success response
            return Response({
                'user': UserSerializer(authenticated_user).data,
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'redirect': '/dashboard/'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            print(f"Login error: {str(e)}")  # Debug log
            return Response(
                {'error': 'An error occurred during login'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RegisterView(generics.CreateAPIView):
    permission_classes = (AllowAny,)
    serializer_class = RegisterSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        refresh = RefreshToken.for_user(user)
        return Response({
            'user': UserSerializer(user).data,
            'refresh': str(refresh),
            'access': str(refresh.access_token),
        }, status=status.HTTP_201_CREATED)

class UserProfileView(generics.RetrieveUpdateAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = UserSerializer

    def get_object(self):
        return self.request.user

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

class PagePermissionsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        permissions = user.get_page_permissions()
        return Response(permissions)

# Data Views
@login_required
def data_entry(request):
    if not request.user.can_view_data_entry:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Data Entry',
        model_name='DataEntry',
        request=request
    )
    
    context = {
        'can_create_data_entry': request.user.can_create_data_entry
    }
    return render(request, 'dashboard/data/data_entry.html', context)

@login_required
def data_edit(request):
    if not request.user.can_view_data_edit:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Data Edit',
        model_name='DataEdit',
        request=request
    )
    
    return render(request, 'dashboard/data/data_edit.html')

@login_required
def enquiry(request):
    if not request.user.can_view_enquiry:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Enquiry',
        model_name='Enquiry',
        request=request
    )
    
    return render(request, 'dashboard/data/enquiry.html')

# Setup Views
@login_required
def branch_dep_link(request):
    if not request.user.can_view_branch_dep_link:
        return render(request, 'dashboard/403.html')
    return render(request, 'dashboard/setup/branch_dep_link.html')

@login_required
def bulk_upload(request):
    if not request.user.can_view_bulk_upload:
        return render(request, 'dashboard/403.html')
    return render(request, 'dashboard/setup/bulk_upload.html')

@login_required
def department(request):
    if not request.user.can_view_department:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Department',
        model_name='Department',
        request=request
    )
    
    context = {
        'can_create_department': request.user.can_create_department,
        'can_update_department': request.user.can_update_department,
        'can_delete_department': request.user.can_delete_department
    }
    return render(request, 'dashboard/setup/department.html', context)

@login_required
def division_branch(request):
    if not request.user.can_view_division_branch:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Division Branch',
        model_name='DivisionBranch',
        request=request
    )
    
    context = {
        'can_create_division_branch': request.user.can_create_division_branch,
        'can_update_division_branch': request.user.can_update_division_branch,
        'can_delete_division_branch': request.user.can_delete_division_branch,
        'can_view_division_branch': request.user.can_view_division_branch
    }
    return render(request, 'dashboard/setup/division_branch.html', context)

@login_required
def logo_upload(request):
    if not request.user.can_view_logo_upload:
        return render(request, 'dashboard/403.html')
    
    # Add debug logging
    print(f"User permissions for logo upload:")
    print(f"can_create_logo_upload: {request.user.can_create_logo_upload}")
    print(f"can_view_logo_upload: {request.user.can_view_logo_upload}")
    print(f"can_update_logo_upload: {request.user.can_update_logo_upload}")
    print(f"can_delete_logo_upload: {request.user.can_delete_logo_upload}")
    
    # Generate token for the user
    refresh = RefreshToken.for_user(request.user)
    token = str(refresh.access_token)
    
    context = {
        'user': request.user,
        'can_create_logo_upload': request.user.can_create_logo_upload,
        'can_view_logo_upload': request.user.can_view_logo_upload,
        'can_update_logo_upload': request.user.can_update_logo_upload,
        'can_delete_logo_upload': request.user.can_delete_logo_upload,
        'token': token  # Add token to context
    }
    return render(request, 'dashboard/setup/logo_upload.html', context)

@login_required
def sub_department(request):
    if not request.user.can_view_sub_department:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Sub Department',
        model_name='SubDepartment',
        request=request
    )
    
    context = {
        'can_create_sub_department': request.user.can_create_sub_department,
        'can_update_sub_department': request.user.can_update_sub_department,
        'can_delete_sub_department': request.user.can_delete_sub_department
    }
    return render(request, 'dashboard/setup/sub_department.html', context)

# User Views
@login_required
def password_change(request):
    """Password change page view"""
    if not request.user.can_access_password_change:
        raise PermissionDenied("You don't have permission to access the password change page.")
    
    # Get all users for the dropdown
    users = User.objects.all().order_by('username')
    
    context = {
        'users': users
    }
    return render(request, 'dashboard/user/password_change.html', context)

@login_required
def change_password(request):
    """API endpoint for changing user password"""
    if not request.user.can_access_password_change:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        new_password = data.get('new_password')
        
        if not user_id or not new_password:
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        # Get the user
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)
        
        # Change the password
        user.set_password(new_password)
        user.save()
        
        # Log the activity
        log_activity(
            user=request.user,
            action='update',
            page='Password Change',
            model_name='User',
            request=request,
            details=f'Changed password for user {user.username}'
        )
        
        return JsonResponse({'message': 'Password changed successfully'})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"Error changing password: {str(e)}")  # Debug log
        return JsonResponse({'error': 'Failed to change password'}, status=500)

@login_required
def users(request):
    """Users management page view"""
    # Check if user has access to users page
    if not request.user.can_access_users:
        raise PermissionDenied("You don't have permission to access the users page.")
    
    # Check if user has view permission
    if not request.user.can_view_users:
        raise PermissionDenied("You don't have permission to view users.")
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Users',
        model_name='User',
        request=request
    )
    
    # Pass permissions to template
    context = {
        'user': request.user,
        'can_create_users': request.user.can_create_users,
        'can_update_users': request.user.can_update_users,
        'can_delete_users': request.user.can_delete_users
    }
    return render(request, 'dashboard/user/users.html', context)

@login_required
def user_list(request):
    """API endpoint for listing and creating users"""
    if request.method == 'GET':
        if not request.user.can_view_users:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        try:
            users = User.objects.all().order_by('-date_joined')
            serializer = UserSerializer(users, many=True)
            return JsonResponse(serializer.data, safe=False)
        except Exception as e:
            print(f"Error fetching users: {str(e)}")  # Debug log
            return JsonResponse({'error': 'Failed to fetch users'}, status=500)
    
    elif request.method == 'POST':
        if not request.user.can_create_users:  # Use the new create permission
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        try:
            data = json.loads(request.body)
            # Check if username or email already exists
            if User.objects.filter(username=data.get('username')).exists():
                return JsonResponse({'error': 'Username already exists'}, status=400)
            if User.objects.filter(email=data.get('email')).exists():
                return JsonResponse({'error': 'Email already exists'}, status=400)
            
            # Create user
            user = User.objects.create_user(
                username=data.get('username'),
                email=data.get('email'),
                password=data.get('password'),
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                status=data.get('status', 'active')
            )
            
            serializer = UserSerializer(user)
            return JsonResponse(serializer.data, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            print(f"Error creating user: {str(e)}")  # Debug log
            return JsonResponse({'error': str(e)}, status=400)

@login_required
def user_detail(request, pk):
    """API endpoint for retrieving, updating and deleting a user"""
    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)

    if request.method == 'GET':
        if not request.user.can_view_users:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        serializer = UserSerializer(user)
        return JsonResponse(serializer.data)

    elif request.method == 'PUT':
        if not request.user.can_edit_users:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        serializer = UserSerializer(user, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        if not request.user.can_delete_users:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        user.delete()
        return JsonResponse({'message': 'User deleted successfully'}, status=204)

@login_required
def user_rights(request):
    """User rights management page view"""
    print("Accessing user rights page")  # Debug log
    
    # Check if user has access to user rights page
    if not request.user.can_access_user_rights:
        print(f"User {request.user.username} does not have access to user rights")  # Debug log
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='User Rights',
        model_name='UserRights',
        request=request
    )
    
    if request.method == 'POST':
        try:
            user_id = request.POST.get('user_id')
            permissions = request.POST.getlist('permissions[]')
            
            print(f"Received request to update permissions for user {user_id}")  # Debug log
            print(f"Permissions to set: {permissions}")  # Debug log
            
            user = User.objects.get(id=user_id)
            print(f"Found user: {user.username}")  # Debug log
            
            # Clear existing permissions
            user.user_permissions.clear()
            print("Cleared existing user permissions")  # Debug log
            
            # Reset all permission fields to False
            for field in user._meta.fields:
                if field.name.startswith('can_'):
                    setattr(user, field.name, False)
            print("Reset all permission fields to False")  # Debug log
            
            # Define permission hierarchy
            permission_hierarchy = {
                'can_access_data': ['can_access_data_entry', 'can_access_data_edit', 'can_access_enquiry'],
                'can_access_data_entry': ['can_view_data_entry', 'can_create_data_entry', 'can_update_data_entry', 'can_delete_data_entry'],
                'can_access_data_edit': ['can_view_data_edit', 'can_create_data_edit', 'can_update_data_edit', 'can_delete_data_edit'],
                'can_access_enquiry': ['can_view_enquiry', 'can_create_enquiry', 'can_update_enquiry', 'can_delete_enquiry'],
                'can_access_setup': ['can_access_department', 'can_access_sub_department', 'can_access_division_branch', 'can_access_branch_dep_link', 'can_access_logo_upload', 'can_access_bulk_upload'],
                'can_access_department': ['can_view_department', 'can_create_department', 'can_update_department', 'can_delete_department'],
                'can_access_sub_department': ['can_view_sub_department', 'can_create_sub_department', 'can_update_sub_department', 'can_delete_sub_department'],
                'can_access_division_branch': ['can_view_division_branch', 'can_create_division_branch', 'can_update_division_branch', 'can_delete_division_branch'],
                'can_access_branch_dep_link': ['can_view_branch_dep_link', 'can_create_branch_dep_link', 'can_update_branch_dep_link', 'can_delete_branch_dep_link'],
                'can_access_logo_upload': ['can_view_logo_upload', 'can_create_logo_upload', 'can_update_logo_upload', 'can_delete_logo_upload'],
                'can_access_bulk_upload': ['can_view_bulk_upload', 'can_create_bulk_upload', 'can_update_bulk_upload', 'can_delete_bulk_upload'],
                'can_access_user': ['can_access_users', 'can_access_user_rights', 'can_access_password_change'],
                'can_access_users': ['can_view_users', 'can_create_users', 'can_update_users', 'can_delete_users'],
                'can_access_user_rights': ['can_view_user_rights', 'can_create_user_rights', 'can_update_user_rights', 'can_delete_user_rights'],
                'can_access_password_change': ['can_view_password_change', 'can_create_password_change', 'can_update_password_change', 'can_delete_password_change'],
                'can_access_report': ['can_access_log_report', 'can_access_register'],
                'can_access_log_report': ['can_view_log_report', 'can_create_log_report', 'can_update_log_report', 'can_delete_log_report'],
                'can_access_register': ['can_view_register', 'can_create_register', 'can_update_register', 'can_delete_register']
            }
            
            # Add new permissions with hierarchy
            for permission in permissions:
                try:
                    # Set the permission
                    if hasattr(user, permission):
                        setattr(user, permission, True)
                        print(f"Set {permission} to True for user {user.username}")  # Debug log
                    
                    # Set parent permissions
                    for parent, children in permission_hierarchy.items():
                        if permission in children:
                            if hasattr(user, parent):
                                setattr(user, parent, True)
                                print(f"Set parent permission {parent} to True")  # Debug log
                    
                    # Set grandparent permissions if they exist
                    for grandparent, parent_list in permission_hierarchy.items():
                        if parent in parent_list:
                            if hasattr(user, grandparent):
                                setattr(user, grandparent, True)
                                print(f"Set grandparent permission {grandparent} to True")  # Debug log
                    
                    # Add to user_permissions if it exists in the Permission model
                    try:
                        perm = Permission.objects.get(codename=permission)
                        user.user_permissions.add(perm)
                        print(f"Added {permission} to user_permissions")  # Debug log
                    except Permission.DoesNotExist:
                        print(f"Permission not found in Permission model: {permission}")
                        continue
                        
                except Exception as e:
                    print(f"Error processing permission {permission}: {str(e)}")
                    continue
            
            # Save the user to update the permission fields
            user.save()
            print(f"Saved user {user.username} with permissions: {[f.name for f in user._meta.fields if f.name.startswith('can_') and getattr(user, f.name)]}")  # Debug log
            
            return JsonResponse({'status': 'success'})
        except User.DoesNotExist:
            print(f"User not found with id: {user_id}")  # Debug log
            return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)
        except Exception as e:
            print(f"Error saving permissions: {str(e)}")  # Debug log
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    # For GET request, render the template
    users = User.objects.all()
    context = {
        'users': users,
        'can_access_users': request.user.can_access_users,
        'can_view_users': request.user.can_view_users
    }
    print(f"Rendering user rights page for user {request.user.username}")  # Debug log
    return render(request, 'dashboard/user/user_rights.html', context)

@login_required
def get_user_permissions(request, user_id):
    """Get permissions for a specific user"""
    try:
        print(f"Fetching permissions for user_id: {user_id}")  # Debug log
        
        # Check if requesting user has permission to view user rights
        if not request.user.can_access_user_rights:
            print(f"User {request.user.username} does not have access to user rights")  # Debug log
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        user = User.objects.get(id=user_id)
        print(f"Found user: {user.username}")  # Debug log
        
        permissions = []
        
        # Get all field-based permissions
        for field in user._meta.fields:
            if field.name.startswith('can_'):
                # Check if the field is a permission field
                if getattr(user, field.name):
                    permissions.append(field.name)
                    print(f"Found permission {field.name} for user {user.username}")  # Debug log
        
        # Get all model-based permissions
        for perm in user.user_permissions.all():
            if perm.codename not in permissions:
                permissions.append(perm.codename)
                print(f"Found model permission {perm.codename}")  # Debug log
        
        # Get all group-based permissions
        for group in user.groups.all():
            for perm in group.permissions.all():
                if perm.codename not in permissions:
                    permissions.append(perm.codename)
                    print(f"Found group permission {perm.codename}")  # Debug log
        
        print(f"Returning permissions: {permissions}")  # Debug log
        return JsonResponse({'permissions': permissions})
    except User.DoesNotExist:
        print(f"User not found with id: {user_id}")  # Debug log
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        print(f"Error getting permissions: {str(e)}")  # Debug log
        return JsonResponse({'error': str(e)}, status=500)

# Report Views
@login_required
def log_report(request):
    if not request.user.can_view_log_report:
        return render(request, 'dashboard/403.html')
    
    # Log the view action
    log_activity(
        user=request.user,
        action='view',
        page='Log Report',
        model_name='ActivityLog',
        request=request
    )
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    user_id = request.GET.get('user')
    action = request.GET.get('action')
    page = request.GET.get('page')
    model = request.GET.get('model')
    download = request.GET.get('download')
    
    # Get activity logs with filters
    logs = ActivityLog.objects.all()
    
    if start_date:
        # Convert to datetime with time set to start of day in IST
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
        start_datetime = timezone.make_aware(start_datetime, timezone=timezone.get_fixed_timezone(330))  # IST is UTC+5:30
        logs = logs.filter(created_at__gte=start_datetime)
    if end_date:
        # Convert to datetime with time set to end of day in IST
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        end_datetime = timezone.make_aware(end_datetime, timezone=timezone.get_fixed_timezone(330))  # IST is UTC+5:30
        logs = logs.filter(created_at__lte=end_datetime)
    if user_id:
        logs = logs.filter(user_id=user_id)
    if action:
        logs = logs.filter(action=action)
    if page:
        logs = logs.filter(page__icontains=page)
    if model:
        logs = logs.filter(model_name__icontains=model)
    
    # Get all users for filter dropdown
    users = User.objects.all()
    
    # Get unique pages and models for filter dropdowns (using distinct to avoid duplicates)
    pages = ActivityLog.objects.values_list('page', flat=True).distinct().order_by('page')
    models = ActivityLog.objects.values_list('model_name', flat=True).distinct().order_by('model_name')
    
    # Handle Excel download
    if download == 'excel':
        # Log the download action
        log_activity(
            user=request.user,
            action='download',
            page='Log Report',
            model_name='ActivityLog',
            details={
                'file_type': 'Excel',
                'file_name': 'activity_logs.xlsx',
                'filters': {
                    'start_date': start_date,
                    'end_date': end_date,
                    'user_id': user_id,
                    'action': action,
                    'page': page,
                    'model': model
                }
            },
            request=request
        )
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activity Logs"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Define headers
        headers = [
            'Timestamp (IST)', 'User', 'Action', 'Page', 'Model', 
            'Object ID', 'Details', 'IP Address', 'User Agent'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, log in enumerate(logs.order_by('-created_at'), 2):
            ws.cell(row=row, column=1, value=log.created_at.astimezone(timezone.get_fixed_timezone(330)).strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=2, value=log.user.get_full_name() if log.user else 'System')
            ws.cell(row=row, column=3, value=log.action)
            ws.cell(row=row, column=4, value=log.page)
            ws.cell(row=row, column=5, value=log.model_name)
            ws.cell(row=row, column=6, value=log.object_id)
            ws.cell(row=row, column=7, value=json.dumps(log.details, indent=2))
            ws.cell(row=row, column=8, value=log.ip_address)
            ws.cell(row=row, column=9, value=log.user_agent)
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Create the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=activity_logs.xlsx'
        
        # Save the workbook to the response
        wb.save(response)
        return response
    
    context = {
        'logs': logs.order_by('-created_at'),
        'users': users,
        'pages': pages,
        'models': models,
        'start_date': start_date,
        'end_date': end_date,
        'selected_user': user_id,
        'selected_action': action,
        'selected_page': page,
        'selected_model': model,
    }
    return render(request, 'dashboard/report/log_report.html', context)

@api_view(['GET'])
@permission_required('dashboard.view_users')
def users_api(request):
    users = User.objects.all()
    serializer = UserSerializer(users, many=True, context={'request': request})
    return Response(serializer.data)

class DepartmentViewSet(mixins.ListModelMixin,
                       mixins.CreateModelMixin,
                       mixins.DestroyModelMixin,
                       mixins.UpdateModelMixin,
                       mixins.RetrieveModelMixin,
                       viewsets.GenericViewSet):
    """
    ViewSet for handling department operations.
    Provides list, create, update, and delete functionality.
    """
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'

    def get_queryset(self):
        if not self.request.user.can_view_department:
            raise PermissionDenied("You don't have permission to view departments")
        return Department.objects.all()

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.request.query_params.get('page', 1)
        per_page = self.request.query_params.get('per_page', 10)
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        serializer = self.get_serializer(page_obj, many=True)
        return Response({
            'departments': serializer.data,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number
        })

    def perform_create(self, serializer):
        if not self.request.user.can_create_department:
            raise PermissionDenied("You don't have permission to create departments")
        # Attach request to the instance
        instance = serializer.save()
        instance._request = self.request
        return instance

    def perform_update(self, serializer):
        if not self.request.user.can_update_department:
            raise PermissionDenied("You don't have permission to update departments")
        # Attach request to the instance
        instance = serializer.save()
        instance._request = self.request
        return instance

    def perform_destroy(self, instance):
        if not self.request.user.can_delete_department:
            raise PermissionDenied("You don't have permission to delete departments")
        # Attach request to the instance before deletion
        instance._request = self.request
        instance.delete()

class SubDepartmentViewSet(mixins.ListModelMixin,
                          mixins.CreateModelMixin,
                          mixins.DestroyModelMixin,
                          mixins.UpdateModelMixin,
                          mixins.RetrieveModelMixin,
                          viewsets.GenericViewSet):
    """
    ViewSet for handling sub-department operations.
    Provides list, create, update, and delete functionality.
    """
    queryset = SubDepartment.objects.all()
    serializer_class = SubDepartmentSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'sub_department_id'

    def get_queryset(self):
        if not self.request.user.can_view_sub_department:
            raise PermissionDenied("You don't have permission to view sub-departments")
        queryset = SubDepartment.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(sub_department_id__icontains=search) |
                Q(department__name__icontains=search)
            )
        return queryset

    def get_object(self):
        try:
            return super().get_object()
        except Http404:
            raise Http404("Sub-department not found")

    def perform_create(self, serializer):
        if not self.request.user.can_create_sub_department:
            raise PermissionDenied("You don't have permission to create sub-departments")
        # Attach request to the instance
        instance = serializer.save()
        instance._request = self.request
        return instance

    def perform_update(self, serializer):
        if not self.request.user.can_update_sub_department:
            raise PermissionDenied("You don't have permission to update sub-departments")
        # Attach request to the instance
        instance = serializer.save()
        instance._request = self.request
        return instance

    def perform_destroy(self, instance):
        if not self.request.user.can_delete_sub_department:
            raise PermissionDenied("You don't have permission to delete sub-departments")
        # Attach request to the instance before deletion
        instance._request = self.request
        instance.delete()

class DivisionBranchViewSet(viewsets.ModelViewSet):
    queryset = DivisionBranch.objects.all()
    serializer_class = DivisionBranchSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        if not self.request.user.can_view_division_branch:
            raise PermissionDenied("You don't have permission to view division/branches")
        queryset = DivisionBranch.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(division_id__icontains=search)
            )
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        if not self.request.user.can_create_division_branch:
            raise PermissionDenied("You don't have permission to create division/branches")
        # Attach request to the instance
        instance = serializer.save()
        instance._request = self.request
        return instance

    def perform_update(self, serializer):
        if not self.request.user.can_update_division_branch:
            raise PermissionDenied("You don't have permission to update division/branches")
        # Attach request to the instance
        instance = serializer.save()
        instance._request = self.request
        return instance

    def perform_destroy(self, instance):
        if not self.request.user.can_delete_division_branch:
            raise PermissionDenied("You don't have permission to delete division/branches")
        # Attach request to the instance before deletion
        instance._request = self.request
        instance.delete()

class BranchDepartmentLinkViewSet(viewsets.ModelViewSet):
    queryset = BranchDepartmentLink.objects.all()
    serializer_class = BranchDepartmentLinkSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.can_view_branch_dep_link:
            raise PermissionDenied("You don't have permission to view branch department links")
            
        queryset = BranchDepartmentLink.objects.all()
        branch = self.request.query_params.get('branch', None)
        department = self.request.query_params.get('department', None)
        sub_department = self.request.query_params.get('sub_department', None)
        is_active = self.request.query_params.get('is_active', None)

        if branch:
            queryset = queryset.filter(branch_id=branch)
        if department:
            queryset = queryset.filter(department_id=department)
        if sub_department:
            queryset = queryset.filter(sub_department_id=sub_department)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        return queryset

    def create(self, request, *args, **kwargs):
        if not request.user.can_create_branch_dep_link:
            raise PermissionDenied("You don't have permission to create branch department links")
        
        # Handle bulk creation
        if isinstance(request.data, list):
            serializer = self.get_serializer(data=request.data, many=True)
            serializer.is_valid(raise_exception=True)
            self.perform_bulk_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        # Handle single creation
        return super().create(request, *args, **kwargs)

    def perform_bulk_create(self, serializer):
        serializer.save()

    def perform_create(self, serializer):
        serializer.save()

    def perform_update(self, serializer):
        if not self.request.user.can_update_branch_dep_link:
            raise PermissionDenied("You don't have permission to update branch department links")
        serializer.save()

    def perform_destroy(self, instance):
        if not self.request.user.can_delete_branch_dep_link:
            raise PermissionDenied("You don't have permission to delete branch department links")
        instance.delete()

class LogoViewSet(viewsets.ModelViewSet):
    queryset = Logo.objects.all()
    serializer_class = LogoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.can_view_logo_upload:
            raise PermissionDenied("You don't have permission to view logos")
        return Logo.objects.all().order_by('-created_at')

    def create(self, request, *args, **kwargs):
        print(f"User in create: {request.user}")
        print(f"can_create_logo_upload: {request.user.can_create_logo_upload}")
        print(f"Request data: {request.data}")
        
        if not request.user.is_authenticated:
            raise PermissionDenied("Authentication required")
            
        if not request.user.can_create_logo_upload:
            raise PermissionDenied("You don't have permission to create logos")
            
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        try:
            print("Starting perform_create")
            print(f"Request data: {self.request.data}")
            
            # Get logo data and check its format
            logo_data = self.request.data.get('logo_data')
            print(f"Logo data type: {type(logo_data)}")
            print(f"Logo data length: {len(logo_data) if logo_data else 0}")
            
            if not logo_data or not isinstance(logo_data, str):
                raise serializers.ValidationError("Invalid logo data format")
                
            # Check if it's a valid base64 string
            if not logo_data.startswith('data:image/'):
                raise serializers.ValidationError("Invalid image data format")
                
            # Split the base64 string and get the actual data
            try:
                base64_data = logo_data.split(',')[1]
                print(f"Base64 data length: {len(base64_data)}")
                logo_binary = base64.b64decode(base64_data)
                print(f"Decoded binary data length: {len(logo_binary)}")
            except Exception as e:
                print(f"Error decoding base64: {str(e)}")
                raise serializers.ValidationError(f"Error decoding image data: {str(e)}")
            
            # Save the logo
            serializer.save(
                logo_data=logo_binary,
                created_by=self.request.user,
                updated_by=self.request.user
            )
            print("Logo saved successfully")
            
        except Exception as e:
            print(f"Error in perform_create: {str(e)}")
            raise serializers.ValidationError(f"Error processing logo data: {str(e)}")

    def update(self, request, *args, **kwargs):
        if not request.user.can_update_logo_upload:
            raise PermissionDenied("You don't have permission to update logos")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not request.user.can_delete_logo_upload:
            raise PermissionDenied("You don't have permission to delete logos")
        return super().destroy(request, *args, **kwargs)

@api_view(['GET', 'POST'])
def user_branch_departments(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=404)

    if request.method == 'GET':
        # Get all branches with their departments and subdepartments
        branches = DivisionBranch.objects.all()
        user_mappings = UserSubDepartment.objects.filter(user=user).first()
        mappings = user_mappings.get_mappings() if user_mappings else {}

        response_data = {
            'branches': []
        }

        for branch in branches:
            branch_data = {
                'id': branch.id,
                'name': branch.name,
                'departments': []
            }

            # Get departments linked to this branch
            branch_dept_links = BranchDepartmentLink.objects.filter(branch=branch)
            for link in branch_dept_links:
                dept = link.department
                dept_data = {
                    'id': dept.id,
                    'name': dept.name,
                    'subdepartments': []
                }

                # Get subdepartments for this department
                subdepartments = SubDepartment.objects.filter(department=dept)
                for subdept in subdepartments:
                    # Check if this subdepartment is assigned to the user
                    key = f"{branch.id}_{dept.id}_{subdept.id}"
                    assigned = key in mappings

                    subdept_data = {
                        'id': subdept.id,
                        'name': subdept.name,
                        'assigned': assigned
                    }
                    dept_data['subdepartments'].append(subdept_data)

                branch_data['departments'].append(dept_data)

            response_data['branches'].append(branch_data)

        return Response(response_data)

    elif request.method == 'POST':
        try:
            data = request.data
            subdepartments = data.get('subdepartments', [])

            # Get or create user mapping
            user_mapping, created = UserSubDepartment.objects.get_or_create(user=user)
            
            # Clear existing mappings
            user_mapping.clear_mappings()

            # Add new mappings
            for mapping in subdepartments:
                branch_id = mapping.get('branch_id')
                department_id = mapping.get('department_id')
                subdepartment_id = mapping.get('subdepartment_id')

                if all([branch_id, department_id, subdepartment_id]):
                    user_mapping.add_mapping(branch_id, department_id, subdepartment_id)

            return Response({'message': 'Branch-department mappings updated successfully'})

        except Exception as e:
            return Response({'error': str(e)}, status=400)

class DataEntryViewSet(viewsets.ModelViewSet):
    serializer_class = DataEntryRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.can_view_data_edit:
            raise PermissionDenied("You don't have permission to view data entries")
        return DataEntryRecord.objects.filter(user=self.request.user)

    def retrieve(self, request, *args, **kwargs):
        if not request.user.can_view_data_edit:
            raise PermissionDenied("You don't have permission to view data entries")
        return super().retrieve(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not request.user.can_update_data_edit:
            raise PermissionDenied("You don't have permission to update data entries")
        
        try:
            instance = self.get_object()
            
            # Parse the JSON data from FormData
            data = json.loads(request.data.get('data', '{}'))
            
            # Handle file updates
            remove_file = request.data.get('remove_file') == 'true'
            new_file = request.FILES.get('file')
            
            # Update the record
            serializer = self.get_serializer(instance, data=data, partial=True)
            serializer.is_valid(raise_exception=True)
            record = serializer.save()
            
            # Handle file updates
            if remove_file:
                # Delete existing files
                record.files.all().delete()
            elif new_file:
                # Delete existing files
                record.files.all().delete()
                # Create new file
                DataEntryFile.objects.create(
                    record=record,
                    file_name=new_file.name,
                    file_type=new_file.content_type,
                    file_data=new_file.read(),
                    file_size=new_file.size
                )
            
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def destroy(self, request, *args, **kwargs):
        if not request.user.can_delete_data_edit:
            raise PermissionDenied("You don't have permission to delete data entries")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def search(self, request):
        try:
            if not request.user.can_view_data_edit:
                raise PermissionDenied("You don't have permission to view data entries")

            # Get search parameters
            branch_id = request.query_params.get('division-filter')
            department_id = request.query_params.get('department-filter')
            sub_department_id = request.query_params.get('subdepartment-filter')
            page = int(request.query_params.get('page', 1))
            per_page = int(request.query_params.get('per_page', 10))

            # Start with base queryset
            queryset = DataEntryRecord.objects.filter(user=request.user)

            # Apply filters
            if branch_id:
                queryset = queryset.filter(branch_id=branch_id)
            if department_id:
                queryset = queryset.filter(department_id=department_id)
            if sub_department_id:
                queryset = queryset.filter(sub_department_id=sub_department_id)

            # Apply field value filters
            for key, value in request.query_params.items():
                if key.startswith('field_') and value:
                    field_name = key.replace('field_', '')
                    queryset = queryset.filter(field_values__has_key=field_name)
                    queryset = queryset.filter(**{f'field_values__{field_name}__icontains': value})

            # Get total count before pagination
            total_count = queryset.count()

            # Apply pagination
            start = (page - 1) * per_page
            end = start + per_page
            queryset = queryset[start:end]

            # Serialize results with permissions
            serializer = self.get_serializer(queryset, many=True)
            results = serializer.data
            
            # Add permission flags to each result
            for result in results:
                result['can_view'] = request.user.can_view_data_edit
                result['can_update'] = request.user.can_update_data_edit
                result['can_delete'] = request.user.can_delete_data_edit
            
            return Response({
                'count': total_count,
                'results': results
            })

        except PermissionDenied as e:
            return Response({'error': str(e)}, status=403)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def get_hierarchy(self, request):
        try:
            # Get user's sub-department mappings
            user_mappings = UserSubDepartment.objects.filter(user=request.user).first()
            if not user_mappings:
                return Response({'error': 'No permissions found for user'}, status=403)

            mappings = user_mappings.get_mappings()
            if not mappings:
                return Response({'error': 'No branch-department mappings found for user'}, status=403)

            # Get unique branch IDs from mappings
            branch_ids = set()
            for mapping in mappings.values():
                branch_ids.add(mapping['branch_id'])

            # Get branches with their departments and sub-departments
            branches = DivisionBranch.objects.filter(id__in=branch_ids)
            serializer = BranchSerializer(branches, many=True, context={'request': request})
            return Response(serializer.data)
        except Exception as e:
            print(f"Error in get_hierarchy: {str(e)}")  # Debug log
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def get_subdepartment_fields(self, request):
        try:
            sub_department_id = request.query_params.get('sub_department_id')
            if not sub_department_id:
                return Response({'error': 'Sub department ID is required'}, status=400)

            # Verify user has access to this sub-department
            user_mappings = UserSubDepartment.objects.filter(user=request.user).first()
            if not user_mappings:
                return Response({'error': 'No permissions found for user'}, status=403)

            mappings = user_mappings.get_mappings()
            if not mappings:
                return Response({'error': 'No branch-department mappings found for user'}, status=403)

            # Check if user has access to this sub-department
            has_access = False
            for mapping in mappings.values():
                if str(mapping['subdepartment_id']) == str(sub_department_id):
                    has_access = True
                    break

            if not has_access:
                return Response({'error': 'You do not have access to this sub-department'}, status=403)

            sub_department = SubDepartment.objects.get(id=sub_department_id)
            return Response({'fields': sub_department.fields})
        except SubDepartment.DoesNotExist:
            return Response({'error': 'Sub department not found'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    def perform_create(self, serializer):
        # Verify user has access to the selected branch, department, and sub-department
        branch_id = serializer.validated_data.get('branch').id
        department_id = serializer.validated_data.get('department').id
        sub_department_id = serializer.validated_data.get('sub_department').id

        user_mappings = UserSubDepartment.objects.filter(user=self.request.user).first()
        if not user_mappings:
            raise PermissionDenied("No permissions found for user")

        mappings = user_mappings.get_mappings()
        if not mappings:
            raise PermissionDenied("No branch-department mappings found for user")

        # Check if user has access to this combination
        has_access = False
        for mapping in mappings.values():
            if (str(mapping['branch_id']) == str(branch_id) and
                str(mapping['department_id']) == str(department_id) and
                str(mapping['subdepartment_id']) == str(sub_department_id)):
                has_access = True
                break

        if not has_access:
            raise PermissionDenied("You do not have access to this branch-department-subdepartment combination")

        # Attach request to the instance
        instance = serializer.save(user=self.request.user)
        instance._request = self.request
        serializer.save(user=self.request.user)

    def create(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                # Get the form data
                branch_id = request.data.get('branch')
                department_id = request.data.get('department')
                sub_department_id = request.data.get('sub_department')
                field_values = json.loads(request.data.get('field_values', '{}'))
                
                print("Request FILES:", request.FILES)  # Debug log
                print("Request content type:", request.content_type)  # Debug log
                print("Request data:", request.data)  # Debug log
                
                # Get files from request
                files = request.FILES.getlist('documents')
                print("Files from getlist:", files)  # Debug log
                print("Number of files:", len(files))  # Debug log
                
                print(f"Received data:")  # Debug log
                print(f"Branch ID: {branch_id}")
                print(f"Department ID: {department_id}")
                print(f"Sub Department ID: {sub_department_id}")
                print(f"Field Values: {field_values}")
                print(f"Number of files: {len(files)}")

                # Create the data entry record
                data = {
                    'branch': branch_id,
                    'department': department_id,
                    'sub_department': sub_department_id,
                    'field_values': field_values,
                    'user': request.user.id
                }
                
                print(f"Creating data entry with data: {data}")  # Debug log
                
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                record = serializer.save(user=request.user)

                print(f"Record created with ID: {record.id}")  # Debug log

                # Handle file uploads
                for file in files:
                    print(f"Processing file: {file.name} ({file.size} bytes)")  # Debug log
                    try:
                        file_obj = DataEntryFile.objects.create(
                            record=record,
                            file_name=file.name,
                            file_type=file.content_type,
                            file_data=file.read(),
                            file_size=file.size
                        )
                        print(f"File saved with ID: {file_obj.id}")  # Debug log
                    except Exception as e:
                        print(f"Error saving file {file.name}: {str(e)}")  # Debug log
                        raise

                return Response(serializer.data, status=201)
        except PermissionDenied as e:
            print(f"Permission denied: {str(e)}")  # Debug log
            return Response({'error': str(e)}, status=403)
        except Exception as e:
            print(f"Error in create: {str(e)}")  # Debug log
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['get'])
    def download_file(self, request, pk=None):
        try:
            record = self.get_object()
            if not record.files.exists():
                return Response({'error': 'No files found for this record'}, status=404)

            # Get the first file (or you could add a file_id parameter to get a specific file)
            file_obj = record.files.first()
            
            # Create the response with the file data
            response = HttpResponse(file_obj.file_data, content_type=file_obj.file_type)
            response['Content-Disposition'] = f'attachment; filename="{file_obj.file_name}"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class LogoutView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            # Get the refresh token from the request
            refresh_token = request.data.get('refresh')
            if refresh_token:
                # Blacklist the refresh token
                token = RefreshToken(refresh_token)
                token.blacklist()
            
            # Clear session
            request.session.flush()
            
            return Response({
                'message': 'Successfully logged out'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing activity logs.
    Only allows viewing, not creating/updating/deleting.
    """
    serializer_class = ActivityLogSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        if not self.request.user.can_view_log_report:
            raise PermissionDenied("You don't have permission to view activity logs")
        
        queryset = ActivityLog.objects.all()
        
        # Filter by date range if provided
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        
        # Filter by user if provided
        user_id = self.request.query_params.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by action if provided
        action = self.request.query_params.get('action')
        if action:
            queryset = queryset.filter(action=action)
        
        # Filter by page if provided
        page = self.request.query_params.get('page')
        if page:
            queryset = queryset.filter(page__icontains=page)
        
        # Filter by model if provided
        model = self.request.query_params.get('model')
        if model:
            queryset = queryset.filter(model_name__icontains=model)
        
        return queryset.order_by('-created_at')

@login_required
def download_bulk_upload_template(request):
    if not request.user.can_view_bulk_upload:
        return render(request, 'dashboard/403.html')
    
    # Create a DataFrame with the template structure
    df = pd.DataFrame({
        'Branch ID': [1, 2, 3],
        'Department ID': [101, 102, 103],
        'Sub Department ID': [1001, 1002, 1003],
        'Field:EmployeeID': ['EMP001', 'EMP002', 'EMP003'],
        'Field:JoiningDate': ['2024-01-01', '2024-01-02', '2024-01-03'],
        'File Name': ['offer_001.pdf', 'resume_002.pdf', 'contract_003.pdf']
    })
    
    # Create Excel writer
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Records', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Records']
        
        # Add some styling
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Set up the response
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=bulk_upload_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response

@login_required
def process_bulk_upload(request):
    if not request.user.can_create_bulk_upload:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get the ZIP file from the request
        zip_file = request.FILES.get('zipFile')
        if not zip_file:
            return JsonResponse({'error': 'No ZIP file provided'}, status=400)
        
        # Create a temporary directory to extract files
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save the ZIP file temporarily
            zip_path = os.path.join(temp_dir, 'upload.zip')
            with open(zip_path, 'wb+') as destination:
                for chunk in zip_file.chunks():
                    destination.write(chunk)
            
            # Extract the ZIP file
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Find the Excel file and data directory
            excel_file = None
            data_dir = None
            
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.endswith(('.xlsx', '.xls')):
                        excel_file = os.path.join(root, file)
                if 'data' in dirs:
                    data_dir = os.path.join(root, 'data')
            
            if not excel_file:
                return JsonResponse({'error': 'No Excel file found in the ZIP'}, status=400)
            if not data_dir:
                return JsonResponse({'error': 'No data directory found in the ZIP'}, status=400)
            
            # Read the Excel file
            df = pd.read_excel(excel_file)
            
            # Convert column names to lowercase for case-insensitive matching
            df.columns = [col.lower() for col in df.columns]
            
            # Define required columns (case-insensitive)
            required_columns = ['branch id', 'department id', 'sub department id', 'file name']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return JsonResponse({
                    'error': f'Missing required columns: {", ".join(missing_columns)}'
                }, status=400)
            
            # Process each row
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Get the file from data directory
                    file_name = row['file name']
                    file_path = os.path.join(data_dir, file_name)
                    
                    if not os.path.exists(file_path):
                        raise FileNotFoundError(f'File not found: {file_name}')
                    
                    # Get branch, department, and sub-department
                    try:
                        branch = DivisionBranch.objects.get(division_id=row['branch id'])
                        department = Department.objects.get(department_id=row['department id'])
                        sub_department = SubDepartment.objects.get(sub_department_id=row['sub department id'])
                    except (DivisionBranch.DoesNotExist, Department.DoesNotExist, SubDepartment.DoesNotExist) as e:
                        raise ValueError(f'Invalid ID in row {index + 2}: {str(e)}')
                    
                    # Get existing fields from sub-department
                    existing_fields = sub_department.fields
                    field_names = {field['name'].lower(): field for field in existing_fields}
                    
                    # Create field values dictionary from dynamic fields
                    field_values = {}
                    new_fields = []
                    
                    for column in df.columns:
                        if column.lower().startswith('field:'):
                            field_name = column.replace('field:', '').strip()
                            field_value = row[column]
                            
                            # Add to field values
                            field_values[field_name] = field_value
                            
                            # Check if field exists in sub-department
                            if field_name.lower() not in field_names:
                                # Create new field configuration
                                new_field = {
                                    'name': field_name,
                                    'data_type': 'alphanumeric',  # Default to alphanumeric
                                    'requirement': 'optional',    # Default to optional
                                    'verify': False              # Default to no verification
                                }
                                new_fields.append(new_field)
                                field_names[field_name.lower()] = new_field
                    
                    # Update sub-department fields if new fields were added
                    if new_fields:
                        sub_department.fields = existing_fields + new_fields
                        sub_department.save()
                    
                    # Create data entry record
                    with transaction.atomic():
                        # Create the record
                        record = DataEntryRecord.objects.create(
                            user=request.user,
                            branch=branch,
                            department=department,
                            sub_department=sub_department,
                            field_values=field_values
                        )
                        
                        # Add the file
                        with open(file_path, 'rb') as f:
                            file_data = f.read()
                            DataEntryFile.objects.create(
                                record=record,
                                file_name=file_name,
                                file_type=mimetypes.guess_type(file_name)[0] or 'application/octet-stream',
                                file_data=file_data,
                                file_size=len(file_data)
                            )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {index + 2}: {str(e)}')
            
            # Log the activity
            log_activity(
                user=request.user,
                action='create',
                page='Bulk Upload',
                model_name='DataEntryRecord',
                details={
                    'success_count': success_count,
                    'error_count': error_count,
                    'errors': errors
                },
                request=request
            )
            
            return JsonResponse({
                'message': 'Bulk upload completed',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            })
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
